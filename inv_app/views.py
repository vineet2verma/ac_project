from datetime import datetime, date
from bson import ObjectId
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from utils.mongo import *
from django.http import Http404
from collections import defaultdict
import math


# Inventory Templates
def inventory_template_download(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Template"

    ws.append([
        "item_name",
        "category",
        "location",
        "unit",
        "opening_qty",
        "rate",
        "reorder_level"
    ])

    # Sample row
    ws.append([
        "Blade 10 mm",
        "Cutting",
        "Godown-1",
        "PCS",
        100,
        25,
        20
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="inventory_bulk_template.xlsx"'
    )

    wb.save(response)
    return response

# Inventory Bulk Upload
def inventory_bulk_upload(request):
    if request.method != "POST":
        return redirect("inv_app:inventory_master")

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        return redirect("inv_app:inventory_master")

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    inv_col = get_inventory_master_collection()
    cat_col = category_collection()

    # Load valid categories
    valid_categories = {
        c["category_name"].strip().lower()
        for c in cat_col.find({"is_active": True})
    }

    added, skipped = 0, 0
    invalid_rows = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            item_name, category, location, unit, opening_qty, rate, reorder_level = row

            if not item_name or not category:
                skipped += 1
                invalid_rows.append(f"Row {index}: Missing item name or category")
                continue

            item_name_clean = item_name.strip().lower()
            category_clean = str(category).strip().lower()

            # Category validation
            if category_clean not in valid_categories:
                skipped += 1
                invalid_rows.append(
                    f"Row {index}: Invalid Category → {category}"
                )
                continue

            # Duplicate check
            exists = inv_col.find_one({
                "item_name": item_name_clean,
                "category": category_clean,
                "is_active": True
            })
            if exists:
                skipped += 1
                invalid_rows.append(f"Row {index}: Duplicate item")
                continue

            opening_qty = float(opening_qty) if opening_qty else 0
            rate = float(rate) if rate else 0
            reorder_level = float(reorder_level) if reorder_level else 0

            inv_col.insert_one({
                "item_name": item_name_clean,
                "category": category_clean,
                "location": location,
                "unit": unit,
                "opening_qty": opening_qty,
                "current_qty": opening_qty,
                "rate": rate,
                "reorder_level": reorder_level,
                "is_active": True,
                "created_at": datetime.now()
            })

            added += 1

        except Exception as e:
            skipped += 1
            invalid_rows.append(f"Row {index}: {str(e)}")

    # Store result in session
    request.session["bulk_upload_summary"] = {
        "added": added,
        "skipped": skipped
    }
    request.session["bulk_upload_errors"] = invalid_rows

    return redirect("inv_app:inventory_master")

# Inventory Add Item in Master
def inventory_master_add(request):
    if request.method == "POST":
        inv_col = get_inventory_master_collection()
        cat_col = category_collection()

        item_name = request.POST.get("item_name", "").strip()
        category = request.POST.get("category", "").strip()

        # 🔴 Validate category against master
        valid_categories = {
            c["category_name"].strip().lower()
            for c in cat_col.find({"is_active": True})
        }

        if category.lower() not in valid_categories:
            request.session["bulk_upload_errors"] = [
                f"Invalid Category → {category} (not found in Category Master)"
            ]
            return redirect("inv_app:inventory_master")

        opening_qty = float(request.POST.get("opening_qty", 0))
        rate = float(request.POST.get("rate", 0))
        reorder_level = float(request.POST.get("reorder_level", 0))

        # 🔴 Duplicate check
        exists = inv_col.find_one({
            "item_name": item_name.lower(),
            "category": category.lower(),
            "is_active": True
        })
        if exists:
            request.session["bulk_upload_errors"] = [
                f"Duplicate Item → {item_name}"
            ]
            return redirect("inv_app:inventory_master")

        inv_col.insert_one({
            "item_name": item_name.lower(),
            "category": category.lower(),
            "location": request.POST.get("location"),
            "unit": request.POST.get("unit"),
            "opening_qty": opening_qty,
            "current_qty": opening_qty,
            "rate": rate,
            "reorder_level": reorder_level,
            "is_active": True,
            "created_at": datetime.now()
        })

        request.session["bulk_upload_summary"] = {
            "added": 1,
            "skipped": 0
        }

    return redirect("inv_app:inventory_master")

# ================= INVENTORY MASTER =================
def inventory_master_view(request):
    inv_col = get_inventory_master_collection()
    cat_col = category_collection()

    # ================= READ & CLEAR BULK UPLOAD SESSION =================
    bulk_summary = request.session.pop("bulk_upload_summary", None)
    bulk_errors = request.session.pop("bulk_upload_errors", None)

    # ================= SAFE FLOAT =================
    def to_float(val):
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    # ================= ADD / UPDATE =================
    if request.method == "POST":
        item_id = request.POST.get("item_id")

        opening_qty = to_float(request.POST.get("opening_qty"))
        rate = to_float(request.POST.get("rate"))
        reorder_level = to_float(request.POST.get("reorder_level"))

        base_data = {
            "item_name": request.POST.get("item_name", "").strip(),
            "category": request.POST.get("category") or None,
            "location": request.POST.get("location", "").strip(),
            "unit": request.POST.get("unit"),
            "rate": rate,
            "reorder_level": reorder_level,
            "is_active": True,
            "updated_at": datetime.now()
        }

        if item_id:
            inv_col.update_one(
                {"_id": ObjectId(item_id)},
                {"$set": base_data}
            )
        else:
            base_data.update({
                "opening_qty": opening_qty,
                "current_qty": opening_qty,
                "created_at": datetime.now()
            })
            inv_col.insert_one(base_data)

        return redirect("inv_app:inventory_master")

    # ================= FILTER PARAMS =================
    search = request.GET.get("search", "").strip()
    category_filter = request.GET.get("category", "")
    location_filter = request.GET.get("location", "")
    low_stock = request.GET.get("low_stock")

    # ================= QUERY BUILD =================
    query = {"is_active": True}

    if search:
        query["item_name"] = {"$regex": search, "$options": "i"}

    if category_filter:
        query["category"] = category_filter

    if location_filter:
        query["location"] = {"$regex": location_filter, "$options": "i"}

    if low_stock:
        query["$expr"] = {"$lte": ["$current_qty", "$reorder_level"]}

    # ================= PAGINATION PARAMS =================
    page = int(request.GET.get("page", 1))
    per_page = int(request.GET.get("per_page", 50))
    page_sizes = [50, 75, 100, 200, 250 ]


    # allow only valid values
    if per_page not in [50, 75, 100, 200, 250]:
        per_page = 50

    skip = (page - 1) * per_page



    # ================= INVENTORY LIST =================
    # items = list(inv_col.find(query).sort("created_at", -1))
    total_count = inv_col.count_documents(query)

    total_pages = math.ceil(total_count / per_page)
    page_range = list(range(1, total_pages + 1))

    items = list(
        inv_col.find(query)
        .sort("created_at", -1)
        .skip(skip)
        .limit(per_page)
    )

    for i in items:
        i["id"] = str(i["_id"])
        i.setdefault("current_qty", 0)
        i.setdefault("rate", 0)
        i.setdefault("reorder_level", 0)

    total_pages = math.ceil(total_count / per_page)

    # ================= CATEGORY LIST =================
    categories = list(cat_col.find({"is_active": True}))
    for c in categories:
        c["id"] = str(c["_id"])

    # ================= STOCK SUMMARY =================
    total_items = len(items)
    total_qty = sum(i.get("current_qty", 0) for i in items)
    total_value = sum(
        i.get("current_qty", 0) * i.get("rate", 0)
        for i in items
    )

    # ================= CATEGORY-WISE SUMMARY =================
    category_summary = []
    grouped = defaultdict(list)

    for item in items:
        category = item.get("category") or "Uncategorized"
        grouped[category].append(item)

    for category, cat_items in grouped.items():
        category_summary.append({
            "category": category,
            "total_items": len(cat_items),
            "total_qty": round(sum(i.get("current_qty", 0) for i in cat_items), 2),
            "total_value": round(
                sum(i.get("current_qty", 0) * i.get("rate", 0) for i in cat_items), 2
            ),
        })



    # ================= RENDER =================
    return render(request, "inv_app/inventory_master.html", {
        "items": items,
        "categories": categories,
        "search": search,
        "category_filter": category_filter,
        "location_filter": location_filter,
        "low_stock": low_stock,

        # 🔥 SUMMARY
        "total_items": total_items,
        "total_qty": round(total_qty, 2),
        "total_value": round(total_value, 2),

        # 🔥 ACCORDION DATA
        "category_summary": category_summary,

        # 🔥 BULK UPLOAD MESSAGES
        "bulk_summary": bulk_summary,
        "bulk_errors": bulk_errors,

        # 🔥 PAGINATION
        "page_sizes": page_sizes,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "total_count": total_count,
        "page_range": page_range,
    })


# Low Stock Alert
def low_stock_alert(request):
    inv_col = get_inventory_master_collection()

    # 🔔 LOW STOCK CONDITION
    items = list(inv_col.find({
        "is_active": True,
        "$expr": {"$lte": ["$current_qty", "$reorder_level"]}
    }).sort("current_qty", 1))

    for i in items:
        i["id"] = str(i["_id"])

    return render(request, "inv_app/low_stock.html", {
        "items": items
    })

def auto_pr_from_low_stock(request):
    if request.method != "POST":
        return redirect("inv_app:low_stock")

    # 🔒 Checkbox check
    if not request.POST.get("auto_pr"):
        return redirect("inv_app:low_stock")

    inv_col = get_inventory_master_collection()
    pr_col = get_purchase_requisition_collection()

    low_items = list(inv_col.find({
        "is_active": True,
        "$expr": {"$lte": ["$current_qty", "$reorder_level"]}
    }))

    created = 0

    for item in low_items:

        # 🔴 Prevent duplicate PR
        exists = pr_col.find_one({
            "item_id": item["_id"],
            "status": {"$in": ["PR_CREATED"]}
        })
        if exists:
            continue

        required_qty = item["reorder_level"] - item["current_qty"]
        if required_qty <= 0:
            continue

        pr_col.insert_one({
            "item_id": item["_id"],
            "item_name": item["item_name"],
            "required_qty": required_qty,
            "status": "PR_CREATED",
            "source": "LOW_STOCK",
            "created_at": datetime.now()
        })

        created += 1

    print(f"Auto PR created: {created}")
    return redirect("inv_app:low_stock")

def create_pr_selected(request):
    if request.method != "POST":
        return redirect("inv_app:low_stock")

    item_ids = request.POST.getlist("item_ids")
    if not item_ids:
        return redirect("inv_app:low_stock")

    inv_col = get_inventory_master_collection()
    pr_col = get_purchase_requisition_collection()

    for item_id in item_ids:
        item = inv_col.find_one({"_id": ObjectId(item_id)})
        if not item:
            continue

        required_qty = item["reorder_level"] - item["current_qty"]
        if required_qty <= 0:
            continue

        # 🔴 Prevent duplicate PR
        exists = pr_col.find_one({
            "item_id": item["_id"],
            "status": "PR_CREATED"
        })
        if exists:
            continue

        pr_col.insert_one({
            "item_id": item["_id"],
            "item_name": item["item_name"],
            "required_qty": required_qty,
            "status": "PR_CREATED",
            "source": "LOW_STOCK",
            "created_at": datetime.now()
        })

    return redirect("inv_app:low_stock")

def download_low_stock_excel(request):
    if request.method != "POST":
        return redirect("inv_app:low_stock")

    item_ids = request.POST.getlist("item_ids")
    if not item_ids:
        return redirect("inv_app:low_stock")

    inv_col = get_inventory_master_collection()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Low Stock Items"

    ws.append([
        "Item Name",
        "Unit",
        "Current Stock",
        "Reorder Level",
        "Required Qty"
    ])

    for item_id in item_ids:
        try:
            item = inv_col.find_one({"_id": ObjectId(item_id)})
        except Exception:
            continue

        if not item:
            continue

        current = float(item.get("current_qty", 0))
        reorder = float(item.get("reorder_level", 0))
        required_qty = max(reorder - current, 0)

        ws.append([
            item.get("item_name", ""),
            item.get("unit", ""),
            current,
            reorder,
            required_qty
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="low_stock_selected.xlsx"'
    )

    wb.save(response)
    return response

# Inventory Delete From Master
def inventory_master_delete(request, pk):
    col = get_inventory_master_collection()
    col.delete_one({"_id": ObjectId(pk)})
    return redirect("inv_app:inventory_master")

# Inventory Ledger Stock In
def inventory_ledger_view(request):
    inv_col = get_inventory_master_collection()
    ledger_col = get_inventory_ledger_collection()

    # ================= ITEMS FOR DROPDOWN =================
    items = list(inv_col.find({"is_active": True}))
    for i in items:
        i["id"] = str(i["_id"])

    # ================= STOCK IN =================
    if request.method == "POST":
        item_id = request.POST.get("item_id")
        qty = float(request.POST.get("qty"))
        rate = float(request.POST.get("rate", 0))
        remarks = request.POST.get("remarks", "")

        item = inv_col.find_one({"_id": ObjectId(item_id)})
        if not item:
            raise Exception("Inventory item not found")

        ledger_col.insert_one({
            "item_id": item["_id"],
            "item_name": item["item_name"],
            "category": item.get("category"),
            "location": item.get("location"),
            "qty": qty,
            "rate": rate,
            "amount": qty * rate,
            "txn_type": "IN",
            "source": "STOCK_IN",
            "remarks": remarks,
            "created_by": request.session.get("mongo_username"),
            "created_at": datetime.now()
        })

        inv_col.update_one(
            {"_id": item["_id"]},
            {"$inc": {"current_qty": qty}}
        )

        return redirect("inv_app:inventory_ledger")

    # ================= FILTER PARAMS =================
    search = request.GET.get("search", "").strip()
    item_filter = request.GET.get("item_id", "")
    txn_type = request.GET.get("txn_type", "")
    remarks_filter = request.GET.get("remarks", "")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    # ================= QUERY BUILD =================
    query = {}

    if search:
        query["item_name"] = {"$regex": search, "$options": "i"}

    if item_filter:
        query["item_id"] = ObjectId(item_filter)

    if txn_type:
        query["txn_type"] = txn_type

    if remarks_filter:
        query["remarks"] = {"$regex": remarks_filter, "$options": "i"}

    if from_date or to_date:
        query["created_at"] = {}
        if from_date:
            query["created_at"]["$gte"] = datetime.fromisoformat(from_date)
        if to_date:
            query["created_at"]["$lte"] = datetime.fromisoformat(to_date)

    # ================= LEDGER LIST =================
    ledger = list(ledger_col.find(query).sort("created_at", -1))
    for l in ledger:
        l["id"] = str(l["_id"])

    # ================= SUMMARY =================
    total_entries = len(ledger)
    total_qty = sum(l.get("qty", 0) for l in ledger)
    total_amount = sum(l.get("amount", 0) for l in ledger)

    return render(request, "inv_app/inventory_ledger.html", {
        "items": items,
        "ledger": ledger,

        # filters
        "search": search,
        "item_filter": item_filter,
        "txn_type": txn_type,
        "remarks_filter": remarks_filter,
        "from_date": from_date,
        "to_date": to_date,

        # summary
        "total_entries": total_entries,
        "total_qty": round(total_qty, 2),
        "total_amount": round(total_amount, 2),
    })

# ✅ Add Inventory Requirement to Order (ERP Correct)
def add_order_inventory(request, order_id):
    inv_master_col = get_inventory_master_collection()
    order_inv_col = get_order_inventory_collection()

    if request.method != "POST":
        return redirect("cnc_work_app:detail", pk=order_id)

    inventory_id = request.POST.get("inventory_id")
    qty = float(request.POST.get("qty", 0))

    if not inventory_id or qty <= 0:
        return redirect("cnc_work_app:detail", pk=order_id)

    inv = inv_master_col.find_one({"_id": ObjectId(inventory_id)})
    if not inv:
        return redirect("cnc_work_app:detail", pk=order_id)

    # ✅ INSERT ERP-CORRECT RECORD
    order_inv_col.insert_one({
        "order_id": ObjectId(order_id),
        "inventory_id": ObjectId(inventory_id),
        "item_name": inv["item_name"],
        "required_qty": qty,          # ✅ CORRECT FIELD
        "reserved_qty": 0,
        "status": "PENDING",
        "rate": float(inv.get("rate", 0)),
        "created_by": request.session.get("mongo_username"),
        "created_at": datetime.now()
    })

    return redirect("cnc_work_app:detail", pk=order_id)


# Delete Inventory From Order (ERP FINAL)
def delete_order_inventory(request, order_id, inv_id):
    order_inv_col = get_order_inventory_collection()

    inv = order_inv_col.find_one({"_id": ObjectId(inv_id)})
    if not inv:
        raise Http404("Inventory record not found")

    if inv.get("status") != "PENDING":
        raise Http404("Cannot delete once inventory is processed")

    order_inv_col.delete_one({"_id": ObjectId(inv_id)})

    return redirect("cnc_work_app:detail", pk=order_id)


# Inventory Stock In Reverse Entry
def delete_stock_in(request, ledger_id):
    inv_col = get_inventory_master_collection()
    ledger_col = get_inventory_ledger_collection()

    if request.method != "POST":
        raise Http404("Invalid request")

    # 🔹 1. FETCH ORIGINAL STOCK IN LEDGER
    ledger = ledger_col.find_one({
        "_id": ObjectId(ledger_id),
        "txn_type": "IN",
        "source": "STOCK_IN"
    })

    if not ledger:
        raise Http404("Stock IN entry not found")

    qty = float(ledger["qty"])
    item_id = ledger["item_id"]

    # 🔹 2. FETCH INVENTORY MASTER
    item = inv_col.find_one({"_id": ObjectId(item_id)})
    if not item:
        raise Http404("Inventory item not found")

    # 🔹 3. CHECK SUFFICIENT STOCK (CRITICAL)
    if item["current_qty"] < qty:
        raise Exception("Cannot delete stock-in: stock already consumed")

    # 🔹 4. LEDGER ENTRY (REVERSAL → OUT)
    ledger_col.insert_one({
        "item_id": item["_id"],
        "item_name": item["item_name"],
        "category": item.get("category"),
        "location": ledger.get("location"),
        "qty": qty,
        "rate": ledger.get("rate", 0),
        "amount": qty * ledger.get("rate", 0),
        "txn_type": "OUT",
        "source": "STOCK_IN_REVERSE",
        "ref_id": ledger["_id"],
        "remarks": "Wrong stock-in entry reversed",
        "created_by": request.session.get("mongo_username"),
        "created_at": datetime.now()
    })

    # 🔹 5. UPDATE INVENTORY MASTER (DECREASE STOCK)
    inv_col.update_one(
        {"_id": item["_id"]},
        {"$inc": {"current_qty": -qty}}
    )

    return redirect("inv_app:inventory_ledger")


def inventory_check(request, order_id):
    order_inv_col = get_order_inventory_collection()
    inv_master_col = get_inventory_master_collection()

    records = list(order_inv_col.find({
        "order_id": ObjectId(order_id)
    }))

    for r in records:
        r["id"] = str(r["_id"])
        r["required_qty"] = float(r.get("required_qty", 0))
        r["reserved_qty"] = float(r.get("reserved_qty", 0))

        item = inv_master_col.find_one({
            "_id": ObjectId(r["inventory_id"])
        })

        r["item_name"] = item.get("item_name") if item else "-"

        # 🔒 1. LOCKED STATES FIRST (MOST IMPORTANT)
        if r.get("status") == "RESERVED":
            r["status_calc"] = "RESERVED"
            r["available_qty"] = "-"
            r["shortage_qty"] = "-"
            continue

        if r.get("status") == "CONSUMED":
            r["status_calc"] = "CONSUMED"
            r["available_qty"] = "-"
            r["shortage_qty"] = "-"
            continue

        # ❌ 2. INVALID ITEM
        if not item:
            r["status_calc"] = "INVALID"
            r["available_qty"] = 0
            r["shortage_qty"] = r["required_qty"]
            continue

        # 📦 3. CHECK STOCK ONLY FOR PENDING
        available = float(item.get("current_qty", 0))
        r["available_qty"] = available

        if available >= r["required_qty"]:
            r["status_calc"] = "AVAILABLE"
            r["shortage_qty"] = "-"
        else:
            r["status_calc"] = "SHORTAGE"
            r["shortage_qty"] = r["required_qty"] - available

    return render(request, "inv_app/inventory_check.html", {
        "records": records,
        "order_id": order_id
    })


def inventory_reserve(request, inv_id):

    order_inv_col = get_order_inventory_collection()
    inv_master_col = get_inventory_master_collection()
    ledger_col = get_inventory_ledger_collection()

    rec = order_inv_col.find_one({"_id": ObjectId(inv_id)})
    if not rec:
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # 🔒 Already reserved or consumed → do nothing
    if rec.get("status") in ["RESERVED", "CONSUMED"]:
        return redirect(request.META.get("HTTP_REFERER", "/"))

    item = inv_master_col.find_one({
        "_id": ObjectId(rec["inventory_id"])
    })
    if not item:
        return redirect(request.META.get("HTTP_REFERER", "/"))

    available = float(item.get("current_qty", 0))
    required = float(rec.get("required_qty", 0))

    # ❌ Not enough stock → do nothing
    if available < required:
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # ================= LOCK STOCK =================
    inv_master_col.update_one(
        {"_id": item["_id"]},
        {"$inc": {"current_qty": -required}}
    )

    order_inv_col.update_one(
        {"_id": rec["_id"]},
        {"$set": {
            "reserved_qty": required,
            "status": "RESERVED",
            "reserved_at": datetime.now()
        }}
    )

    ledger_col.insert_one({
        "item_id": item["_id"],
        "item_name": item["item_name"],
        "order_id": rec["order_id"],
        "qty": required,
        "txn_type": "RESERVE",
        "source": "ORDER",
        "ref_id": rec["_id"],
        "created_at": datetime.now()
    })

    return redirect(request.META.get("HTTP_REFERER", "/"))


def create_purchase_requisition(request, inv_id):
    inv_col = get_order_inventory_collection()
    pr_col = get_purchase_requisition_collection()

    inv = inv_col.find_one({"_id": ObjectId(inv_id)})
    if not inv:
        raise Http404("Order inventory not found")

    # 🔴 Prevent duplicate PR
    exists = pr_col.find_one({
        "order_inventory_id": inv["_id"],
        "status": {"$ne": "CANCELLED"}
    })
    if exists:
        raise Http404("PR already created")

    pr_col.insert_one({
        "order_id": inv["order_id"],
        "order_inventory_id": inv["_id"],
        "item_id": ObjectId(inv["inventory_id"]),   # ✅ FIX
        "item_name": inv["item_name"],               # ✅ OPTIONAL BUT GOOD
        "required_qty": inv["required_qty"],
        "status": "PR_CREATED",
        "created_at": datetime.now()
    })

    inv_col.update_one(
        {"_id": inv["_id"]},
        {"$set": {"status": "PR_CREATED"}}
    )

    return redirect(request.META.get("HTTP_REFERER", "/"))


def pr_list(request, order_id):
    pr_col = get_purchase_requisition_collection()


    prs = list(
        pr_col.find({"order_id": ObjectId(order_id)})
        .sort("created_at", -1)
    )

    for pr in prs:
        pr["id"] = str(pr["_id"])

    return render(request, "inv_app/pr_list.html", {
        "prs": prs,
        "order_id": order_id
    })


def material_received(request, pr_id):
    if request.method != "POST":
        raise Http404("Invalid request")

    pr_col = get_purchase_requisition_collection()
    inv_col = get_inventory_master_collection()
    order_inv_col = get_order_inventory_collection()
    ledger_col = get_inventory_ledger_collection()

    pr = pr_col.find_one({"_id": ObjectId(pr_id)})
    if not pr or pr["status"] != "PR_CREATED":
        raise Http404("Invalid PR")

    received_qty = float(request.POST.get("received_qty", 0))
    if received_qty <= 0:
        raise Http404("Invalid received quantity")

    # 🔹 UPDATE INVENTORY MASTER (FULL QTY)
    inv_col.update_one(
        {"_id": pr["item_id"]},
        {"$inc": {"current_qty": received_qty}}
    )

    # 🔹 LEDGER ENTRY
    ledger_col.insert_one({
        "item_id": pr["item_id"],
        "qty": received_qty,
        "txn_type": "IN",
        "source": "PURCHASE",
        "ref_id": pr["_id"],
        "created_at": datetime.now()
    })

    # 🔹 UPDATE PR
    pr_col.update_one(
        {"_id": pr["_id"]},
        {"$set": {
            "status": "RECEIVED",
            "received_qty": received_qty,
            "received_at": datetime.now()
        }}
    )

    # 🔹 UPDATE ORDER INVENTORY STATUS
    order_inv_col.update_one(
        {"_id": pr["order_inventory_id"]},
        {"$set": {"status": "AVAILABLE"}}
    )

    return redirect("inv_app:pr_list", order_id=str(pr["order_id"]))


def cancel_pr(request, pr_id):

    pr_col = get_purchase_requisition_collection()
    order_inv_col = get_order_inventory_collection()

    pr = pr_col.find_one({"_id": ObjectId(pr_id)})
    if not pr:
        raise Http404("Purchase Requisition not found")

    # 🔒 SAFETY CHECK
    if pr.get("status") != "PR_CREATED":
        raise Http404("Only pending PR can be cancelled")

    # 🔁 UPDATE PR STATUS
    pr_col.update_one(
        {"_id": pr["_id"]},
        {"$set": {
            "status": "CANCELLED",
            "cancelled_at": datetime.now()
        }}
    )

    # 🔁 RESET ORDER INVENTORY
    order_inv_col.update_one(
        {"_id": pr["order_inventory_id"]},
        {"$set": {
            "status": "SHORTAGE"
        }}
    )

    return redirect(request.META.get("HTTP_REFERER"))


# Inventory Category Master
def category_master(request):
    col = category_collection()
    if request.method == "POST":
        name = request.POST.get("category_name")
        if name:
            col.insert_one({
                "category_name": name,
                "is_active": True,
                "created_at": datetime.now()
            })
        return redirect("inv_app:category_master")

    # categories = list(col.find({"is_active": True}))
    categories = list(col.find({"is_active": {"$ne": False}}))
    for c in categories:
        c["id"] = str(c["_id"])

    return render(
        request, "inv_app/category_master.html",
        {"categories": categories}
    )


# Inventory Category Delete
def category_delete(request, pk):
    col = category_collection()
    col.delete_one({"_id": ObjectId(pk)})
    return redirect("inv_app:category_master")
